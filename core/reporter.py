from datetime import datetime
from pathlib import Path
from typing import NamedTuple

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from core.config import API_URL, OUTPUT_EXCEL, OUTPUT_PDF

_PDF_UNICODE_REPLACEMENTS = {
    "–": "-", "—": "-",     # en/em dash
    "‘": "'", "’": "'",     # aspas simples tipográficas
    "“": '"', "”": '"',     # aspas duplas tipográficas
    "…": "...",                  # reticências
    " ": " ",                    # espaço não separável
}


def _sanitize_pdf_text(text) -> str:
    """A fonte core "Helvetica" do PDF só suporta Latin-1/WinAnsi — títulos e
    descrições de editais vêm de fonte externa e às vezes trazem tipografia
    Unicode (travessão, aspas curvas) que derruba o fpdf2 com
    FPDFUnicodeEncodingException. Normaliza os casos comuns e, pra qualquer
    sobra, cai para o caractere mais próximo em vez de quebrar a geração."""
    if text is None:
        return ""
    text = str(text)
    for k, v in _PDF_UNICODE_REPLACEMENTS.items():
        text = text.replace(k, v)
    try:
        text.encode("latin-1")
    except UnicodeEncodeError:
        text = text.encode("latin-1", errors="replace").decode("latin-1")
    return text


class _PDFReport(FPDF):
    def header(self):
        if self.page_no() > 1:
            self.set_font("Helvetica", "B", 8)
            self.set_text_color(0, 51, 102)
            self.cell(0, 5, "Análise de Editais PNUD Brasil", align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_draw_color(0, 51, 102)
            self.line(self.l_margin, self.get_y() + 1, self.l_margin + self.w, self.get_y() + 1)
            self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 7)
        self.set_text_color(140)
        self.cell(0, 8, f"Página {self.page_no()}/{{nb}} | {datetime.now().strftime('%d/%m/%Y')}", align="C")

    def titulo(self, text: str):
        self.ln(2)
        self.set_font("Helvetica", "B", 12)
        self.set_text_color(0, 51, 102)
        self.cell(0, 8, _sanitize_pdf_text(text), new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(0, 51, 102)
        self.line(self.l_margin, self.get_y() + 1, self.l_margin + self.w, self.get_y() + 1)
        self.ln(4)

    def subtitulo(self, text: str):
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(0, 70, 130)
        self.cell(0, 6, _sanitize_pdf_text(text), new_x="LMARGIN", new_y="NEXT")
        self.ln(2)

    def corpo(self, text: str):
        self.set_font("Helvetica", "", 9)
        self.set_text_color(60, 60, 60)
        self.multi_cell(0, 5, _sanitize_pdf_text(text))
        self.ln(1)

    def bullet(self, text: str, indent: int = 5):
        self.set_font("Helvetica", "", 9)
        self.set_text_color(60, 60, 60)
        x0 = self.l_margin + indent
        self.set_x(x0)
        self.cell(3, 5, "-")
        self.set_x(x0 + 4)
        self.multi_cell(self.w - indent - 4, 5, _sanitize_pdf_text(text))
        self.ln(0.5)

    def tabela(self, headers: list, rows: list, col_widths: list | None = None):
        if col_widths is None:
            col_widths = [self.w / len(headers)] * len(headers)
        total_w = sum(col_widths)
        col_widths = [w * self.w / total_w for w in col_widths]

        self.set_font("Helvetica", "B", 7)
        self.set_fill_color(0, 51, 102)
        self.set_text_color(255)
        for i, h in enumerate(headers):
            self.cell(col_widths[i], 6, _sanitize_pdf_text(h), border=1, fill=True)
        self.ln()

        self.set_font("Helvetica", "", 7)
        for r_idx, row in enumerate(rows):
            if r_idx % 2 == 0:
                self.set_fill_color(242, 246, 252)
            else:
                self.set_fill_color(255, 255, 255)
            self.set_text_color(60)
            for i, cell in enumerate(row):
                self.cell(col_widths[i], 5, _sanitize_pdf_text(cell)[:80], border=1, fill=True)
            self.ln()
        self.ln(3)


_AZUL = "003366"
_AZUL_MEDIO = "004682"
_BANDA = "F2F6FC"
_CINZA_BORDA = "D6DEE8"

_FILL_CABECALHO = PatternFill("solid", fgColor=_AZUL)
_FILL_BANDA = PatternFill("solid", fgColor=_BANDA)
_FONTE_CABECALHO = Font(bold=True, color="FFFFFF", size=10)
_FONTE_LINK = Font(color="0563C1", underline="single", size=10)
_BORDA = Border(**{lado: Side(style="thin", color=_CINZA_BORDA)
                   for lado in ("left", "right", "top", "bottom")})

_FMT_MOEDA = 'R$ #,##0.00'
_FMT_DATA = "DD/MM/YYYY"


class _Coluna(NamedTuple):
    titulo: str
    chave: str
    largura: int
    formato: str | None = None
    horizontal: str = "left"
    quebrar: bool = False
    link: bool = False


def _data(valor):
    """Converte 'YYYY-MM-DD' em date para o Excel tratar como data de verdade."""
    try:
        return datetime.strptime(valor, "%Y-%m-%d").date()
    except (TypeError, ValueError):
        return valor or ""


def _montar_aba(wb: Workbook, nome: str, colunas: list[_Coluna], linhas: list[dict]):
    ws = wb.create_sheet(nome)

    for idx, col in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=idx, value=col.titulo)
        celula.fill = _FILL_CABECALHO
        celula.font = _FONTE_CABECALHO
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = _BORDA
        ws.column_dimensions[get_column_letter(idx)].width = col.largura
    ws.row_dimensions[1].height = 26

    for n, item in enumerate(linhas, start=2):
        for idx, col in enumerate(colunas, start=1):
            valor = item.get(col.chave)
            celula = ws.cell(row=n, column=idx, value="" if valor is None else valor)
            celula.border = _BORDA
            celula.alignment = Alignment(horizontal=col.horizontal, vertical="top",
                                         wrap_text=col.quebrar)
            if col.formato:
                celula.number_format = col.formato
            if col.link and valor:
                celula.hyperlink = valor
                celula.font = _FONTE_LINK
            if n % 2 == 0:
                celula.fill = _FILL_BANDA

    ws.freeze_panes = "A2"
    if linhas:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{len(linhas) + 1}"

    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    return ws


def _montar_resumo(wb: Workbook, analise: dict):
    ws = wb.create_sheet("Resumo")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 46
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Análise de Editais PNUD Brasil"
    ws["A1"].font = Font(bold=True, size=16, color=_AZUL)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 24

    gerado = analise.get("data_analise", datetime.now().isoformat())
    ws["A2"] = f"Gerado em {datetime.fromisoformat(gerado).strftime('%d/%m/%Y às %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="8A8A8A")
    ws.merge_cells("A2:B2")

    linha = 4

    def secao(titulo: str):
        nonlocal linha
        celula = ws.cell(row=linha, column=1, value=titulo)
        celula.font = _FONTE_CABECALHO
        celula.fill = _FILL_CABECALHO
        celula.alignment = Alignment(vertical="center")
        ws.cell(row=linha, column=2).fill = _FILL_CABECALHO
        ws.row_dimensions[linha].height = 20
        linha += 1

    def metrica(rotulo: str, valor, formato: str | None = None):
        nonlocal linha
        rot = ws.cell(row=linha, column=1, value=rotulo)
        rot.font = Font(bold=True, size=10, color="404040")
        rot.border = _BORDA
        val = ws.cell(row=linha, column=2, value="—" if valor is None else valor)
        val.border = _BORDA
        val.alignment = Alignment(horizontal="left", wrap_text=True)
        if formato and valor is not None:
            val.number_format = formato
        linha += 1

    filtro = analise.get("filtro_aplicado", {})
    secao("Panorama")
    metrica("Total de editais", analise.get("total_editais", 0))
    if filtro.get("todos"):
        metrica("Período", "Histórico completo")
    elif filtro.get("periodo_meses"):
        metrica("Período", f"Últimos {filtro['periodo_meses']} meses")
    if filtro.get("perfil"):
        metrica("Perfil filtrado", filtro["perfil"])
    metrica("Tipos distintos", len(analise.get("contagem_tipos", {})))
    metrica("Órgãos parceiros", len(analise.get("contagem_orgaos", {})))
    linha += 1

    valores = analise.get("valores") or {}
    if valores.get("quantidade_com_valor"):
        secao("Valores estimados")
        metrica("Editais com valor divulgado", valores["quantidade_com_valor"])
        metrica("Menor valor", valores.get("minimo"), _FMT_MOEDA)
        metrica("Maior valor", valores.get("maximo"), _FMT_MOEDA)
        metrica("Valor médio", valores.get("medio"), _FMT_MOEDA)
        metrica("Valor mediano", valores.get("mediano"), _FMT_MOEDA)
        linha += 1

    contagem_perfis = {k: v for k, v in (analise.get("contagem_perfis") or {}).items() if k}
    if contagem_perfis:
        secao("Editais por perfil")
        for nome, qtd in contagem_perfis.items():
            metrica(nome.replace("_", " ").title(), qtd)

    return ws


def gerar_excel(analise: dict) -> Path:
    editais = analise.get("editais") or []

    wb = Workbook()
    wb.remove(wb.active)

    _montar_resumo(wb, analise)

    linhas = []
    for e in editais:
        areas = e.get("areas_tematicas") or []
        perfil = e.get("perfil_classificado") or ""
        score = (e.get("matches") or {}).get(perfil, {}).get("score", e.get("score_perfil"))
        linhas.append({
            "torid": e.get("torid"),
            "titulo": e.get("titulo", ""),
            "tipo": e.get("tipo", ""),
            "perfil": perfil.replace("_", " ").title(),
            "score": score,
            "areas": ", ".join(areas) if isinstance(areas, list) else areas,
            "orgao_parceiro": e.get("orgao_parceiro", ""),
            "local": e.get("local", ""),
            "data_inicio": _data(e.get("data_inicio")),
            "data_fim": _data(e.get("data_fim")),
            "valor": e.get("valor_estimado_num") or None,
            "status": e.get("status", ""),
            "email_submissao": e.get("email_submissao", ""),
            "url_externo": e.get("url_externo") or API_URL,
        })

    _montar_aba(wb, "Editais", [
        _Coluna("ToR", "torid", 10, "0", "center"),
        _Coluna("Título", "titulo", 52, quebrar=True),
        _Coluna("Tipo", "tipo", 26, quebrar=True),
        _Coluna("Perfil", "perfil", 22),
        _Coluna("Score", "score", 9, "0.00", "center"),
        _Coluna("Áreas temáticas", "areas", 28, quebrar=True),
        _Coluna("Órgão parceiro", "orgao_parceiro", 20, quebrar=True),
        _Coluna("Local", "local", 20, quebrar=True),
        _Coluna("Início", "data_inicio", 12, _FMT_DATA, "center"),
        _Coluna("Prazo final", "data_fim", 12, _FMT_DATA, "center"),
        _Coluna("Valor estimado", "valor", 16, _FMT_MOEDA, "right"),
        _Coluna("Status", "status", 14, horizontal="center"),
        _Coluna("E-mail", "email_submissao", 30),
        _Coluna("Link", "url_externo", 34, link=True),
    ], linhas)

    contagens = [
        ("Por_Perfil", "Perfil", analise.get("contagem_perfis"), True),
        ("Por_Tipo", "Tipo", analise.get("contagem_tipos"), False),
        ("Por_Area", "Área temática", analise.get("contagem_areas"), False),
        ("Por_Orgao", "Órgão parceiro", analise.get("contagem_orgaos"), False),
    ]
    for nome_aba, rotulo, contagem, formatar in contagens:
        if not contagem:
            continue
        itens = [{"chave": (k.replace("_", " ").title() if formatar else k), "qtd": v}
                 for k, v in contagem.items() if k]
        if itens:
            _montar_aba(wb, nome_aba, [
                _Coluna(rotulo, "chave", 40, quebrar=True),
                _Coluna("Quantidade", "qtd", 14, "0", "center"),
            ], itens)

    por_perfil = analise.get("por_perfil") or {}
    if por_perfil:
        itens = [{"perfil": nome.replace("_", " ").title(),
                  "quantidade": d.get("quantidade", 0),
                  "descricao": d.get("descricao", "")}
                 for nome, d in por_perfil.items()]
        _montar_aba(wb, "Perfis", [
            _Coluna("Perfil", "perfil", 24),
            _Coluna("Editais compatíveis", "quantidade", 18, "0", "center"),
            _Coluna("Descrição", "descricao", 70, quebrar=True),
        ], itens)

    wb.save(OUTPUT_EXCEL)
    return OUTPUT_EXCEL


def gerar_pdf(analise: dict) -> Path:
    editais = analise["editais"]
    total = analise["total_editais"]
    filtro = analise.get("filtro_aplicado", {})

    pdf = _PDFReport()
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=18)

    _pdf_capa(pdf, total, filtro)
    _pdf_visao_geral(pdf, analise)
    _pdf_perfis(pdf, analise)
    _pdf_valores(pdf, analise)
    _pdf_lista_editais(pdf, editais)

    pdf.output(str(OUTPUT_PDF))
    return OUTPUT_PDF


def _pdf_capa(pdf: _PDFReport, total: int, filtro: dict):
    pdf.add_page()
    pdf.ln(25)
    pdf.set_font("Helvetica", "B", 24)
    pdf.set_text_color(0, 51, 102)
    pdf.cell(0, 12, "EDITAIS PNUD BRASIL", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 13)
    pdf.cell(0, 8, "Análise de Editais e Classificação por Perfil", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)
    pdf.set_draw_color(0, 51, 102)
    pdf.line(pdf.l_margin + 25, pdf.get_y(), pdf.l_margin + pdf.w - 25, pdf.get_y())
    pdf.ln(8)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(80)
    filtro_desc = "Todos os editais" if filtro.get("todos") else f"Últimos {filtro.get('periodo_meses', 3)} meses"
    pdf.cell(0, 7, f"Período: {filtro_desc}", align="C", new_x="LMARGIN", new_y="NEXT")
    if filtro.get("perfil"):
        pdf.cell(0, 7, f"Perfil: {filtro['perfil']}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 7, f"{total} editais analisados | {datetime.now().strftime('%d/%m/%Y')}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 7, "Fonte: parceiros.undp.org.br/opportunities", align="C", new_x="LMARGIN", new_y="NEXT")


def _pdf_visao_geral(pdf: _PDFReport, analise: dict):
    pdf.add_page()
    pdf.titulo("1. VISÃO GERAL")
    pdf.corpo(f"Total de editais analisados: {analise['total_editais']}")

    if analise.get("contagem_tipos"):
        pdf.subtitulo("1.1 Por Tipo de Edital")
        rows = [[k, str(v)] for k, v in analise["contagem_tipos"].items()]
        pdf.tabela(["Tipo", "Quantidade"], rows, [120, 50])

    if analise.get("contagem_areas"):
        pdf.subtitulo("1.2 Áreas Temáticas")
        rows = [[k, str(v)] for k, v in analise["contagem_areas"].items()]
        pdf.tabela(["Área Temática", "Quantidade"], rows, [120, 50])

    if analise.get("contagem_orgaos"):
        pdf.subtitulo("1.3 Órgãos Parceiros")
        rows = [[k, str(v)] for k, v in analise["contagem_orgaos"].items()]
        pdf.tabela(["Órgão", "Quantidade"], rows, [120, 50])


def _pdf_perfis(pdf: _PDFReport, analise: dict):
    pdf.add_page()
    pdf.titulo("2. CLASSIFICAÇÃO POR PERFIL")
    pdf.corpo("Cada edital foi automaticamente classificado de acordo com o perfil mais compatível, "
              "baseado nas áreas temáticas, ferramentas exigidas, graduações e idiomas.")

    if analise.get("contagem_perfis"):
        rows = [[k, str(v)] for k, v in analise["contagem_perfis"].items() if k != "Não classificado"]
        if rows:
            pdf.tabela(["Perfil", "Quantidade"], rows, [120, 50])

    if analise.get("por_perfil"):
        pdf.subtitulo("2.1 Detalhamento por Perfil")
        for nome, dados in analise["por_perfil"].items():
            pdf.subtitulo(f"Perfil: {nome}")
            pdf.corpo(f"Descrição: {dados.get('descricao', 'N/D')}")
            pdf.corpo(f"Editais compatíveis: {dados['quantidade']}")
            for e in dados.get("editais", [])[:3]:
                score = e.get("score", 0)
                pdf.bullet(f"[ID {e['id']}] {e['titulo'][:120]} (score: {score:.0%})")


def _pdf_valores(pdf: _PDFReport, analise: dict):
    valores = analise.get("valores", {})
    if not valores or valores.get("quantidade_com_valor", 0) == 0:
        return
    pdf.add_page()
    pdf.titulo("3. VALORES ESTIMADOS")
    pdf.corpo(f"Editais com valor identificado: {valores['quantidade_com_valor']}")
    for label, key in [("Mínimo", "minimo"), ("Máximo", "maximo"), ("Médio", "medio"), ("Mediano", "mediano")]:
        v = valores.get(key)
        if v:
            pdf.corpo(f"{label}: R$ {v:,.2f}")


def _pdf_lista_editais(pdf: _PDFReport, editais: list):
    pdf.add_page()
    pdf.titulo("4. LISTA DE EDITAIS")
    for e in editais[:50]:
        titulo = e.get("titulo", "")[:100]
        pdf.bullet(f"[{e.get('id', '')}] {titulo}")
        pdf.bullet(f"Tipo: {e.get('tipo', '')} | Perfil: {e.get('perfil_classificado', '')} "
                   f"| Valor: R$ {e.get('valor_estimado') or 'NI'} "
                   f"| Órgão: {e.get('orgao_parceiro', '')}", indent=8)
        pdf.ln(1)


def gerar_relatorio_completo(analise: dict, novidades: dict | None = None) -> tuple[Path, Path]:
    from core.site_generator import gerar_dados_site, _mesclar_valores_tors

    _mesclar_valores_tors(analise["editais"])

    excel_path = gerar_excel(analise)
    print(f"📊 Excel salvo em: {excel_path}")
    pdf_path = gerar_pdf(analise)
    print(f"📄 PDF salvo em: {pdf_path}")

    site_json, perfis_json = gerar_dados_site(analise, novidades)
    print(f"🌐 Dados do site salvos em: {site_json}")
    print(f"👤 Perfis do site salvos em: {perfis_json}")
    return excel_path, pdf_path
